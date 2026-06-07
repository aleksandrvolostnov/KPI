import io
import locale
import os
from calendar import month_name, monthcalendar
from datetime import datetime, timedelta
from functools import wraps
from hashlib import md5
from collections import defaultdict, Counter

import matplotlib

matplotlib.use('Agg')
import matplotlib.pyplot as plt
import networkx as nx
import tempfile

from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, send_from_directory, url_for
from flask_login import LoginManager, UserMixin, current_user, login_required, login_user, logout_user
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config["SECRET_KEY"] = "your_secret_key"
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL", "postgresql://postgres:1111@localhost/efficiency_control"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"


# ------------------ МОДЕЛИ ------------------
class Role(db.Model):
    __tablename__ = "roles"
    id = db.Column(db.Integer, primary_key=True)
    role_name = db.Column(db.String(50), nullable=False)


class User(db.Model, UserMixin):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey("roles.id"))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    role = db.relationship("Role", backref="users")
    email = db.Column(db.String(150), unique=True, nullable=True)
    phone = db.Column(db.String(20), nullable=True)
    avatar = db.Column(db.String(255), nullable=True)
    tokens = db.Column(db.Integer, default=0)
    sent_messages = db.relationship(
        "Message",
        foreign_keys="Message.sender_id",
        backref="sender_rel",
        lazy="dynamic",
        overlaps="sender_rel,sent_messages_backref",
    )
    received_messages = db.relationship(
        "Message",
        foreign_keys="Message.receiver_id",
        backref="receiver_rel",
        lazy="dynamic",
        overlaps="receiver_rel,received_messages_backref",
    )


class Project(db.Model):
    __tablename__ = "projects"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    owner_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    owner = db.relationship("User", backref="owned_projects", lazy=True)
    tasks = db.relationship("Task", backref="project", lazy="dynamic")
    members = db.relationship("User", secondary="project_members", backref="projects")


project_members = db.Table(
    "project_members",
    db.Column("project_id", db.Integer, db.ForeignKey("projects.id"), primary_key=True),
    db.Column("user_id", db.Integer, db.ForeignKey("users.id"), primary_key=True),
)


class Reminder(db.Model):
    __tablename__ = "reminders"
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text, nullable=True)
    reminder_date = db.Column(db.DateTime, nullable=False)
    priority = db.Column(db.String(50), default="Низкий")
    repeat = db.Column(db.String(50), default="Нет")
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    user = db.relationship("User", backref="reminders")


class Message(db.Model):
    __tablename__ = "messages"
    id = db.Column(db.Integer, primary_key=True)
    sender_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    receiver_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)
    is_group = db.Column(db.Boolean, default=False)
    content = db.Column(db.Text, nullable=True)
    filename = db.Column(db.String(150), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    is_read = db.Column(db.Boolean, default=False)
    parent_message_id = db.Column(db.Integer, db.ForeignKey("messages.id"), nullable=True)
    sender = db.relationship(
        "User", foreign_keys=[sender_id], backref="sent_messages_backref", overlaps="sent_messages,sender_rel"
    )
    receiver = db.relationship(
        "User",
        foreign_keys=[receiver_id],
        backref="received_messages_backref",
        lazy="joined",
        overlaps="received_messages,receiver_rel",
    )
    parent_message = db.relationship("Message", remote_side=[id], backref="replies", lazy="joined")


# ---------------- КАНБАН МОДЕЛИ ----------------
class KanbanColumn(db.Model):
    __tablename__ = "kanban_columns"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    wip_limit = db.Column(db.Integer, default=5)
    order = db.Column(db.Integer, default=0)
    project_id = db.Column(db.Integer, db.ForeignKey("projects.id"), nullable=True)
    project = db.relationship("Project", backref="kanban_columns")

    def to_dict(self):
        return {"id": self.id, "name": self.name, "wip_limit": self.wip_limit, "order": self.order}


class Task(db.Model):
    __tablename__ = "tasks"
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text, nullable=False)
    priority = db.Column(db.String(50), nullable=False)
    status = db.Column(db.String(50), nullable=False, default="To Do")
    difficulty = db.Column(db.String(50), nullable=False)
    due_date = db.Column(db.DateTime, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"))
    assigned_to_id = db.Column(db.Integer, db.ForeignKey("users.id"))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    project_id = db.Column(db.Integer, db.ForeignKey("projects.id"), nullable=True)
    parent_task_id = db.Column(db.Integer, db.ForeignKey("tasks.id"))
    parent_task = db.relationship("Task", remote_side=[id], backref="dependent_tasks")
    user = db.relationship("User", foreign_keys=[user_id], backref="tasks_created")
    assigned_to = db.relationship("User", foreign_keys=[assigned_to_id], backref="tasks_assigned")

    # Канбан поля
    position = db.Column(db.Integer, default=0)
    started_at = db.Column(db.DateTime, nullable=True)
    review_at = db.Column(db.DateTime, nullable=True)
    completed_at = db.Column(db.DateTime, nullable=True)
    status_history = db.Column(db.JSON, default=list)

    def to_dict(self):
        return {
            "id": self.id,
            "title": self.title,
            "description": self.description,
            "priority": self.priority,
            "status": self.status,
            "due_date": self.due_date.strftime("%Y-%m-%d"),
            "assigned_to": self.assigned_to.username if self.assigned_to else None,
            "project": self.project.name if self.project else None,
            "subtasks": [st.to_dict() for st in self.subtasks],
            "parent_task": self.parent_task.title if self.parent_task else None,
            "dependent_tasks": [t.title for t in self.dependent_tasks],
        }


class SubTask(db.Model):
    __tablename__ = "subtasks"
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey("tasks.id"), nullable=False)
    title = db.Column(db.String(150), nullable=False)
    start_date = db.Column(db.DateTime, nullable=False)
    end_date = db.Column(db.DateTime, nullable=False)
    task = db.relationship("Task", backref=db.backref("subtasks", lazy=True))

    def to_dict(self):
        return {
            "id": self.id,
            "task_id": self.task_id,
            "title": self.title,
            "start_date": self.start_date.strftime("%Y-%m-%d") if self.start_date else "",
            "end_date": self.end_date.strftime("%Y-%m-%d") if self.end_date else "",
        }


class TaskComments(db.Model):
    __tablename__ = "task_comments"
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey("tasks.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    comment = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    task = db.relationship("Task", backref="task_comments")
    user = db.relationship("User", backref="user_comments")

    def to_dict(self):
        return {
            "id": self.id,
            "task_id": self.task_id,
            "user_id": self.user_id,
            "comment": self.comment,
            "created_at": self.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        }


class File(db.Model):
    __tablename__ = "files"
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    task_id = db.Column(db.Integer, db.ForeignKey("tasks.id"))


# ---------- НОВАЯ МОДЕЛЬ ДЛЯ ВЕСОВ КРИТЕРИЕВ ----------
class UserWeight(db.Model):
    __tablename__ = 'user_weights'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), unique=True, nullable=False)
    weight_priority = db.Column(db.Float, default=0.5)  # вес приоритета
    weight_difficulty = db.Column(db.Float, default=0.3)  # вес сложности
    weight_time = db.Column(db.Float, default=0.2)  # вес времени в статусе
    ahp_matrix = db.Column(db.JSON, nullable=True)  # сохранённая матрица сравнений

    user = db.relationship('User', backref='weights')


# ------------------ ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ------------------
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def format_difficulty(value):
    """Преобразует числовую сложность в текст для отображения"""
    if value in (1, "1"):
        return "Легкая"
    elif value in (2, "2"):
        return "Средняя"
    elif value in (3, "3"):
        return "Сложная"
    return str(value)


# ================= [PROCESS MINING] ОСНОВНЫЕ ФУНКЦИИ =================
def get_task_event_log(user_id=None, project_id=None):
    """Извлекает логи событий из статус-истории задач."""
    query = Task.query
    if user_id:
        query = query.filter((Task.user_id == user_id) | (Task.assigned_to_id == user_id))
    if project_id:
        query = query.filter_by(project_id=project_id)

    tasks = query.all()
    events = []
    for task in tasks:
        history = task.status_history or []
        events.append({
            'task_id': task.id,
            'title': task.title,
            'timestamp': task.created_at.isoformat(),
            'from_status': None,
            'to_status': task.status,
            'user': task.assigned_to.username if task.assigned_to else None
        })
        last_status = task.status
        for step in history:
            new_status = step['status']
            events.append({
                'task_id': task.id,
                'title': task.title,
                'timestamp': step['timestamp'],
                'from_status': last_status,
                'to_status': new_status,
                'user': task.assigned_to.username if task.assigned_to else None
            })
            last_status = new_status

    events.sort(key=lambda x: x['timestamp'])
    return events


def compute_process_metrics(events):
    """Вычисляет основные метрики: среднее время в статусах, матрицу переходов."""
    status_durations = defaultdict(list)
    transition_counts = Counter()
    tasks_events = defaultdict(list)

    for ev in events:
        tasks_events[ev['task_id']].append(ev)

    for task_id, ev_list in tasks_events.items():
        for i in range(len(ev_list) - 1):
            current = ev_list[i]
            nxt = ev_list[i + 1]
            transition_counts[(current['to_status'], nxt['to_status'])] += 1
            if current['timestamp'] and nxt['timestamp']:
                start = datetime.fromisoformat(current['timestamp'])
                end = datetime.fromisoformat(nxt['timestamp'])
                duration_hours = (end - start).total_seconds() / 3600
                status_durations[current['to_status']].append(duration_hours)

    avg_duration = {}
    for status, durations in status_durations.items():
        avg_duration[status] = round(sum(durations) / len(durations), 2) if durations else 0

    transitions = [{'from': f, 'to': t, 'count': c} for (f, t), c in transition_counts.items() if f != t]

    return {
        'avg_duration_per_status': avg_duration,
        'transitions': transitions,
        'total_tasks': len(tasks_events),
        'total_events': len(events)
    }


def generate_transition_graph(transitions):
    """Строит граф переходов с помощью networkx, возвращает путь к временному PNG."""
    if not transitions:
        return None
    G = nx.DiGraph()
    for t in transitions:
        from_node = t['from'] if t['from'] else 'Создание'
        to_node = t['to']
        G.add_edge(from_node, to_node, weight=t['count'])
    if G.number_of_edges() == 0:
        return None

    plt.figure(figsize=(12, 8), facecolor='#1e1e2a')
    pos = nx.spring_layout(G, k=2, seed=42)
    nx.draw_networkx_nodes(G, pos, node_size=3000, node_color='#4a4a5a', edgecolors='#6ABFB1', linewidths=2)
    weights = [G[u][v]['weight'] for u, v in G.edges()]
    max_weight = max(weights) if weights else 1
    edge_widths = [1 + 5 * w / max_weight for w in weights]
    nx.draw_networkx_edges(G, pos, width=edge_widths, edge_color='#88aaff', arrows=True, arrowsize=20, arrowstyle='->')
    edge_labels = {(u, v): G[u][v]['weight'] for u, v in G.edges()}
    nx.draw_networkx_edge_labels(G, pos, edge_labels=edge_labels, font_size=10, font_color='white')
    nx.draw_networkx_labels(G, pos, font_size=12, font_color='white', font_weight='bold')
    plt.title('Граф переходов между статусами задач', fontsize=16, color='white')
    plt.axis('off')
    plt.tight_layout()

    tmp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
    plt.savefig(tmp_file.name, format='png', facecolor='#1e1e2a', edgecolor='none')
    plt.close()
    return tmp_file.name


def generate_recommendations(metrics, transitions, regress_count):
    """Формирует управленческие рекомендации на основе метрик."""
    recs = []
    avg_duration = metrics['avg_duration_per_status']
    total_tasks = metrics['total_tasks']

    if avg_duration:
        max_status = max(avg_duration.items(), key=lambda x: x[1])
        max_hours = max_status[1]
        avg_hours = sum(avg_duration.values()) / len(avg_duration) if avg_duration else 0
        if max_hours > avg_hours * 2 and max_hours > 8:
            recs.append({
                'type': 'bottleneck',
                'title': 'Узкое место в потоке',
                'message': f'Статус «{max_status[0]}» занимает {max_hours:.1f} ч, что в {max_hours / avg_hours:.1f} раз больше среднего.',
                'action': 'Проведите анализ причин задержек в этом статусе, рассмотрите WIP-лимиты или дополнительное обучение.'
            })

    if regress_count > 0 and total_tasks > 0:
        regress_rate = regress_count / total_tasks
        if regress_rate > 0.3:
            recs.append({
                'type': 'quality',
                'title': 'Высокий процент возвратов',
                'message': f'{regress_count} задач ({regress_rate * 100:.0f}%) вернулись из Review -> In Progress.',
                'action': 'Введите чек-лист для ревью, назначьте ответственного за качество, проводите ретроспективы.'
            })
        elif regress_rate > 0.1:
            recs.append({
                'type': 'quality',
                'title': 'Есть возвраты',
                'message': f'{regress_count} задач потребовали доработки после ревью.',
                'action': 'Уточните критерии приёмки задачи и проводите короткие встречи по результатам ревью.'
            })

    if 'In Progress' in avg_duration and avg_duration['In Progress'] > 24:
        recs.append({
            'type': 'long_cycle',
            'title': 'Задачи застревают в работе',
            'message': f'Среднее время в статусе In Progress: {avg_duration["In Progress"]:.1f} ч (> 1 дня).',
            'action': 'Разбивайте задачи на более мелкие, используйте ежедневные стендапы для выявления блокеров.'
        })
    if 'Review' in avg_duration and avg_duration['Review'] > 12:
        recs.append({
            'type': 'review_slow',
            'title': 'Медленное ревью',
            'message': f'Ревью длится в среднем {avg_duration["Review"]:.1f} ч.',
            'action': 'Установите тайм-лимит на ревью (2-4 часа), внедрите парное программирование или автоматические проверки.'
        })

    if total_tasks > 0 and len(transitions) == 0:
        recs.append({
            'type': 'no_flow',
            'title': 'Процесс не движется',
            'message': 'Нет переходов между статусами. Возможно, задачи никогда не меняют статус.',
            'action': 'Настройте доску так, чтобы каждая задача обязательно проходила этапы. Проведите обучение команды работе с канбаном.'
        })

    if regress_count == 0 and total_tasks > 5 and len(avg_duration) >= 3:
        recs.append({
            'type': 'good',
            'title': 'Отличный поток!',
            'message': 'Нет возвратов на доработку, средние времена в норме.',
            'action': 'Документируйте текущие практики и делитесь опытом с другими командами.'
        })
    return recs


# ================= РАСШИРЕННЫЕ МЕТРИКИ =================
def calculate_extended_metrics(events):
    """Lead Time, Cycle Time, Throughput, статистика по исполнителям."""
    tasks_data = defaultdict(lambda: {'created': None, 'completed': None, 'status_timeline': []})
    for ev in events:
        tid = ev['task_id']
        if ev['from_status'] is None:
            tasks_data[tid]['created'] = datetime.fromisoformat(ev['timestamp'])
        tasks_data[tid]['status_timeline'].append(ev)

    lead_times = []
    cycle_times = []
    completed_tasks = 0
    performer_stats = defaultdict(lambda: {'total_time': 0, 'task_count': 0})

    for tid, data in tasks_data.items():
        if data['created'] is None:
            continue
        last_done = None
        active_time = 0.0
        for i, ev in enumerate(data['status_timeline']):
            if ev['to_status'] == 'Done':
                last_done = datetime.fromisoformat(ev['timestamp'])
            if ev['to_status'] in ('In Progress', 'Review') and i + 1 < len(data['status_timeline']):
                start = datetime.fromisoformat(ev['timestamp'])
                end = datetime.fromisoformat(data['status_timeline'][i + 1]['timestamp'])
                active_time += (end - start).total_seconds() / 3600
        if last_done:
            lead = (last_done - data['created']).total_seconds() / 3600
            lead_times.append(lead)
            cycle_times.append(active_time)
            completed_tasks += 1
            assignee = data['status_timeline'][-1].get('user')
            if assignee:
                performer_stats[assignee]['total_time'] += lead
                performer_stats[assignee]['task_count'] += 1

    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    throughput_count = sum(
        1 for ev in events if ev['to_status'] == 'Done' and datetime.fromisoformat(ev['timestamp']) >= thirty_days_ago)
    throughput = round(throughput_count / 30, 2) if completed_tasks else 0

    return {
        'avg_lead_time_hours': round(sum(lead_times) / len(lead_times), 2) if lead_times else 0,
        'avg_cycle_time_hours': round(sum(cycle_times) / len(cycle_times), 2) if cycle_times else 0,
        'throughput_per_day': throughput,
        'completed_tasks': completed_tasks,
        'performer_stats': [
            {'name': name,
             'avg_time_hours': round(stats['total_time'] / stats['task_count'], 2) if stats['task_count'] else 0,
             'count': stats['task_count']}
            for name, stats in performer_stats.items()
        ]
    }


def get_timeline_events(events):
    """Возвращает список событий для тепловой карты (дата, статус)."""
    timeline = []
    for ev in events:
        dt = datetime.fromisoformat(ev['timestamp']).date()
        timeline.append({'date': dt.isoformat(), 'status': ev['to_status']})
    return timeline


def get_sankey_data(transitions):
    """Формирует данные для Sankey-диаграммы (узлы и связи)."""
    nodes = set()
    links = []
    for t in transitions:
        from_node = t['from'] if t['from'] else 'Создание'
        to_node = t['to']
        nodes.add(from_node)
        nodes.add(to_node)
        links.append({'source': from_node, 'target': to_node, 'value': t['count']})
    nodes_list = list(nodes)
    source_indices = [nodes_list.index(link['source']) for link in links]
    target_indices = [nodes_list.index(link['target']) for link in links]
    return {
        'nodes': nodes_list,
        'links': links,
        'source_indices': source_indices,
        'target_indices': target_indices,
        'values': [l['value'] for l in links]
    }


# ================= НОРМАЛИЗАЦИЯ И ПОЛЕЗНОСТЬ (методы из учебника) =================
def normalize_minimization(values):
    """Min-max нормализация для критерия, который нужно минимизировать (чем меньше, тем лучше)."""
    if not values:
        return []
    min_val = min(values)
    max_val = max(values)
    if max_val == min_val:
        return [1.0] * len(values)
    return [(max_val - v) / (max_val - min_val) for v in values]


def normalize_maximization(values):
    """Min-max нормализация для критерия, который нужно максимизировать."""
    if not values:
        return []
    min_val = min(values)
    max_val = max(values)
    if max_val == min_val:
        return [1.0] * len(values)
    return [(v - min_val) / (max_val - min_val) for v in values]


def get_numeric_priority(priority_str):
    mapping = {'Низкий': 1, 'Средний': 2, 'Высокий': 3}
    return mapping.get(priority_str, 2)


def get_numeric_difficulty(difficulty_str):
    if difficulty_str in ('1', 'Легко', 'Легкая'):
        return 1
    if difficulty_str in ('2', 'Средне', 'Средняя'):
        return 2
    if difficulty_str in ('3', 'Сложно', 'Сложная'):
        return 3
    return 2


def get_user_weights(user_id):
    """Возвращает объект UserWeight для пользователя, создаёт со значениями по умолчанию, если нет."""
    weights = UserWeight.query.filter_by(user_id=user_id).first()
    if not weights:
        weights = UserWeight(user_id=user_id, weight_priority=0.5, weight_difficulty=0.3, weight_time=0.2)
        db.session.add(weights)
        db.session.commit()
    return weights


def compute_normalized_utilities(tasks, user_weights):
    """Для списка задач возвращает список кортежей (task, utility, norm_priority, norm_difficulty, norm_time)."""
    if not tasks:
        return []
    priorities = [get_numeric_priority(t.priority) for t in tasks]
    difficulties = [get_numeric_difficulty(t.difficulty) for t in tasks]
    times = []
    for t in tasks:
        if t.status in ('In Progress', 'Review') and t.started_at:
            times.append((datetime.utcnow() - t.started_at).total_seconds() / 3600)
        else:
            times.append(0)

    norm_priorities = normalize_maximization(priorities)
    norm_difficulties = normalize_minimization(difficulties)
    norm_times = normalize_maximization(times)

    result = []
    for i, task in enumerate(tasks):
        utility = (user_weights.weight_priority * norm_priorities[i] +
                   user_weights.weight_difficulty * norm_difficulties[i] +
                   user_weights.weight_time * norm_times[i])
        result.append((task, utility, norm_priorities[i], norm_difficulties[i], norm_times[i]))
    return result


def ahp_weights_from_matrix(matrix):
    """
    Реализация метода аналитической иерархии (Саати) для матрицы 3x3.
    Возвращает (weights, consistency_ratio).
    """
    import numpy as np
    n = len(matrix)
    m = np.array(matrix, dtype=float)
    # Суммируем по столбцам
    col_sums = m.sum(axis=0)
    # Нормируем матрицу
    norm_matrix = m / col_sums
    # Усредняем по строкам – это веса
    weights = np.mean(norm_matrix, axis=1)
    # Вычисление λ_max для проверки согласованности
    lambda_max = np.mean(np.sum(m, axis=1) * weights) / np.sum(weights)
    ci = (lambda_max - n) / (n - 1)
    ri = {1: 0, 2: 0, 3: 0.58, 4: 0.9, 5: 1.12}.get(n, 1.24)
    cr = ci / ri if ri != 0 else 0
    return weights.tolist(), cr


# ------------------ МАРШРУТЫ (ОСНОВНЫЕ) ------------------
@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    if request.method == "POST":
        if "email" in request.form:
            email = request.form["email"]
            if email:
                current_user.email = email
                flash("Почта обновлена!", "success")
            else:
                current_user.email = None
                flash("Почта удалена!", "success")
        if "phone" in request.form:
            phone = request.form["phone"]
            if phone:
                current_user.phone = phone
                flash("Телефон обновлен!", "success")
            else:
                current_user.phone = None
                flash("Телефон удален!", "success")
        if "avatar" in request.files:
            avatar_file = request.files["avatar"]
            if avatar_file and allowed_file(avatar_file.filename):
                avatar_filename = secure_filename(avatar_file.filename)
                avatar_path = os.path.join(app.config["UPLOAD_FOLDER"], avatar_filename)
                avatar_file.save(avatar_path)
                current_user.avatar = avatar_filename
                flash("Аватарка обновлена!", "success")
        db.session.commit()
        return redirect(url_for("profile"))
    return render_template("profile.html", user=current_user)


@app.route("/user/<int:user_id>")
@login_required
def view_user(user_id):
    user = User.query.get_or_404(user_id)
    return render_template("user_profile.html", user=user)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.route("/")
def index():
    return render_template("welcome.html")


# ------------------ ЧАТ ------------------
def mark_messages_as_read(sender_id, receiver_id):
    unread_messages = Message.query.filter_by(receiver_id=receiver_id, sender_id=sender_id, is_read=False).all()
    for msg in unread_messages:
        msg.is_read = True
        db.session.commit()


def get_last_message(user_id_1, user_id_2):
    return (
        Message.query.filter(
            ((Message.sender_id == user_id_1) & (Message.receiver_id == user_id_2))
            | ((Message.sender_id == user_id_2) & (Message.receiver_id == user_id_1))
        )
        .order_by(Message.created_at.desc())
        .first()
    )


@app.route("/chat/<user_id>", methods=["GET", "POST"])
@login_required
def chat(user_id):
    users = User.query.filter(User.id != current_user.id).all()
    last_messages = {}
    user_has_new_message = {}
    for user in users:
        last_message = get_last_message(current_user.id, user.id)
        last_messages[user.id] = (
            last_message.content
            if last_message and last_message.content
            else (f"Файл: {last_message.filename}" if last_message else "Нет сообщений")
        )
        has_new_message = Message.query.filter_by(sender_id=user.id, receiver_id=current_user.id,
                                                  is_read=False).count() > 0
        user_has_new_message[user.id] = has_new_message
    users_sorted = sorted(
        users,
        key=lambda user: (
            get_last_message(current_user.id, user.id).created_at
            if get_last_message(current_user.id, user.id)
            else datetime.min
        ),
        reverse=True,
    )
    if request.method == "POST":
        content = request.form.get("content")
        file = request.files.get("file")
        parent_message_id = request.form.get("parent_message_id")
        if not content and not file:
            return jsonify({"status": "error", "message": "Сообщение не может быть пустым"})
        filename = None
        if file:
            filename = secure_filename(file.filename)
            file.save(os.path.join("uploads", filename))
        new_message = Message(sender_id=current_user.id, receiver_id=user_id, content=content, filename=filename)
        if parent_message_id:
            new_message.parent_message_id = parent_message_id
        db.session.add(new_message)
        db.session.commit()
        return jsonify({"status": "success", "message": "Сообщение отправлено"})
    if user_id == "group":
        messages = Message.query.filter_by(is_group=True).order_by(Message.created_at.asc()).all()
    else:
        messages = (
            Message.query.filter(
                ((Message.sender_id == current_user.id) & (Message.receiver_id == user_id))
                | ((Message.sender_id == user_id) & (Message.receiver_id == current_user.id))
            )
            .order_by(Message.created_at.asc())
            .options(db.joinedload(Message.parent_message))
            .all()
        )
    unread = Message.query.filter_by(receiver_id=current_user.id, sender_id=user_id, is_read=False).all()
    for msg in unread:
        msg.is_read = True
    db.session.commit()
    chat_with = "Общая группа" if user_id == "group" else User.query.get(user_id).username
    return render_template(
        "chat.html",
        messages=messages,
        users=users_sorted,
        chat_with=chat_with,
        last_messages=last_messages,
        user_has_new_message=user_has_new_message,
    )


@app.route("/chat/new_message", methods=["POST"])
@login_required
def new_message():
    content = request.form.get("content")
    receiver_id = request.form.get("receiver_id")
    parent_message_id = request.form.get("parent_message_id")
    if content and receiver_id:
        new_message = Message(sender_id=current_user.id, receiver_id=receiver_id, content=content)
        if parent_message_id:
            new_message.parent_message_id = parent_message_id
        db.session.add(new_message)
        db.session.commit()
        return jsonify(
            {
                "status": "success",
                "message": "Сообщение отправлено",
                "new_message": {"sender": current_user.username, "content": content},
            }
        )
    return jsonify({"status": "error", "message": "Невозможно отправить сообщение"})


@app.route("/forward", methods=["POST"])
@login_required
def forward_message():
    message_id = request.form.get("message_id")
    recipient_id = request.form.get("recipient_id")
    if not message_id or not recipient_id:
        return jsonify({"status": "error", "message": "Неверные данные для пересылки"}), 400
    message = Message.query.get(message_id)
    if not message:
        return jsonify({"status": "error", "message": "Сообщение не найдено"}), 404
    forwarded = Message(
        sender_id=current_user.id,
        receiver_id=recipient_id,
        content=f"Пересланное сообщение: {message.content}",
        filename=message.filename,
    )
    db.session.add(forwarded)
    db.session.commit()
    return jsonify({"status": "success", "message": "Сообщение переслано"})


@app.route("/reply/<message_id>", methods=["POST"])
@login_required
def reply_to_message(message_id):
    content = request.form.get("reply_content")
    original = Message.query.get(message_id)
    if original and content:
        reply = Message(
            sender_id=current_user.id, receiver_id=original.sender_id, content=content, parent_message_id=original.id
        )
        db.session.add(reply)
        db.session.commit()
        return jsonify(
            {
                "status": "success",
                "message": "Ответ отправлен",
                "new_message": {"sender": current_user.username, "content": content},
            }
        )
    return jsonify({"status": "error", "message": "Невозможно отправить ответ"})


@app.route("/uploads/<filename>")
@login_required
def download_file(filename):
    return send_from_directory("uploads", filename)


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        confirm = request.form["confirm_password"]
        role_id = request.form["role_id"]
        if password != confirm:
            flash("Пароли не совпадают!", "error")
            return redirect(url_for("register"))
        if User.query.filter_by(username=username).first():
            flash("Пользователь с таким именем уже существует!", "error")
            return redirect(url_for("register"))
        new_user = User(username=username, password=password, role_id=role_id)
        db.session.add(new_user)
        db.session.commit()
        flash("Регистрация прошла успешно!", "success")
        return redirect(url_for("login"))
    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        user = User.query.filter_by(username=username, password=password).first()
        if user:
            login_user(user)
            return redirect(url_for("dashboard"))
        else:
            flash("Неверное имя пользователя или пароль!")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("index"))


# ------------------ ДАШБОРД ------------------
@app.route("/dashboard")
@login_required
def dashboard():
    if current_user.role.role_name == "Admin":
        tasks = Task.query.all()
    else:
        tasks = Task.query.filter((Task.user_id == current_user.id) | (Task.assigned_to_id == current_user.id)).all()
    tasks_sorted = sorted(tasks, key=lambda x: x.due_date)
    upcoming_tasks = tasks_sorted[:3]
    total_tasks = len(tasks)
    low_priority_count = Task.query.filter_by(priority="Низкий").count()
    medium_priority_count = Task.query.filter_by(priority="Средний").count()
    high_priority_count = Task.query.filter_by(priority="Высокий").count()
    easy_count = Task.query.filter_by(difficulty="1").count() + Task.query.filter_by(difficulty="Легко").count()
    medium_count = Task.query.filter_by(difficulty="2").count() + Task.query.filter_by(difficulty="Средне").count()
    hard_count = Task.query.filter_by(difficulty="3").count() + Task.query.filter_by(difficulty="Сложно").count()
    avg_difficulty = db.session.query(db.func.avg(db.cast(Task.difficulty, db.Float))).scalar()
    return render_template(
        "dashboard.html",
        tasks=upcoming_tasks,
        total_tasks=total_tasks,
        low_priority_count=low_priority_count,
        medium_priority_count=medium_priority_count,
        high_priority_count=high_priority_count,
        average_difficulty=avg_difficulty,
        easy_count=easy_count,
        medium_count=medium_count,
        hard_count=hard_count,
    )


# ------------------ НАСТРОЙКИ ЗАГРУЗКИ ФАЙЛОВ ------------------
UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploads")
ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "gif", "docx", "xlsx", "xls", "doc", "csv"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ------------------ СОЗДАНИЕ И РЕДАКТИРОВАНИЕ ЗАДАЧ ------------------
@app.route("/task/new", methods=["GET", "POST"])
@login_required
def create_task():
    project_id = request.args.get("project_id")
    project = None
    if project_id:
        project = Project.query.get_or_404(project_id)
        if current_user.id != project.owner_id and current_user not in project.members:
            flash("У вас нет доступа для добавления задач.", "error")
            return redirect(url_for("projects"))
    if request.method == "POST":
        title = request.form["title"]
        description = request.form["description"]
        due_date = datetime.strptime(request.form["due_date"], "%Y-%m-%d")
        difficulty = request.form["difficulty"]
        priority = request.form["priority"]
        status = request.form["status"]
        assigned_to_id = request.form.get("assigned_to")
        new_task = Task(
            title=title,
            description=description,
            due_date=due_date,
            difficulty=difficulty,
            priority=priority,
            status=status,
            user_id=current_user.id,
            assigned_to_id=assigned_to_id,
            project_id=project.id if project else None,
            position=0,
        )
        db.session.add(new_task)
        files = request.files.getlist("task_files[]")
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
                db.session.add(File(filename=filename, task_id=new_task.id))
        db.session.commit()
        if project:
            return redirect(url_for("project_detail", project_id=project.id))
        else:
            return redirect(url_for("dashboard"))
    if project:
        columns = (
            KanbanColumn.query.filter((KanbanColumn.project_id == project.id) | (KanbanColumn.project_id.is_(None)))
            .order_by(KanbanColumn.order)
            .all()
        )
    else:
        columns = KanbanColumn.query.filter_by(project_id=None).order_by(KanbanColumn.order).all()
    status_choices = [(col.name, col.name) for col in columns]
    users = project.members if project else User.query.all()
    return render_template("create_task.html", users=users, project=project, status_choices=status_choices)


@app.route("/task/delete/<int:task_id>", methods=["POST"])
@login_required
def delete_task(task_id):
    task = Task.query.get_or_404(task_id)
    if task.user != current_user and current_user.role.role_name != "Admin":
        return redirect(url_for("dashboard"))
    db.session.delete(task)
    db.session.commit()
    return redirect(url_for("dashboard"))


@app.route("/delete_file/<int:file_id>", methods=["POST"])
@login_required
def delete_file(file_id):
    file = File.query.get_or_404(file_id)
    path = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
    if os.path.exists(path):
        os.remove(path)
    db.session.delete(file)
    db.session.commit()
    return jsonify({"status": "success"})


@app.route("/task/edit/<int:task_id>", methods=["GET", "POST"])
@login_required
def edit_task(task_id):
    task = Task.query.get_or_404(task_id)
    if request.method == "POST":
        task.title = request.form["title"]
        task.description = request.form["description"]
        task.due_date = datetime.strptime(request.form["due_date"], "%Y-%m-%d")
        task.difficulty = request.form["difficulty"]
        task.priority = request.form["priority"]
        task.status = request.form["status"]
        task.assigned_to_id = request.form.get("assigned_to")
        files = request.files.getlist("task_files[]")
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
                db.session.add(File(filename=filename, task_id=task.id))
        db.session.commit()
        return redirect(url_for("dashboard"))
    files = File.query.filter_by(task_id=task_id).all()
    users = User.query.all()
    columns = (
        KanbanColumn.query.filter((KanbanColumn.project_id == task.project_id) | (KanbanColumn.project_id.is_(None)))
        .order_by(KanbanColumn.order)
        .all()
    )
    status_choices = [(col.name, col.name) for col in columns]
    return render_template("edit_task.html", task=task, users=users, files=files, status_choices=status_choices)


@app.route("/tasks", methods=["GET", "POST"])
@login_required
def tasks():
    users = User.query.all()
    all_statuses = [col.name for col in KanbanColumn.query.all()]
    all_statuses = list(dict.fromkeys(all_statuses))
    if request.method == "POST":
        query = Task.query
        if request.form.get("title"):
            query = query.filter(Task.title.ilike(f"%{request.form['title']}%"))
        if request.form.get("assigned_to"):
            query = query.filter(Task.assigned_to_id == int(request.form["assigned_to"]))
        if request.form.get("priority"):
            query = query.filter(Task.priority == request.form["priority"])
        if request.form.get("due_date"):
            query = query.filter(Task.due_date == datetime.strptime(request.form["due_date"], "%Y-%m-%d"))
        if request.form.get("difficulty"):
            query = query.filter(Task.difficulty == request.form["difficulty"])
        if request.form.get("status"):
            query = query.filter(Task.status == request.form["status"])
        tasks = query.all()
    else:
        tasks = Task.query.all()
    return render_template("tasks.html", tasks=tasks, users=users, all_statuses=all_statuses)


# ------------------ КАЛЕНДАРЬ ------------------
try:
    locale.setlocale(locale.LC_TIME, "ru_RU.UTF-8")
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, "ru_RU")
    except locale.Error:
        locale.setlocale(locale.LC_TIME, "en_US.UTF-8")


@app.route("/calendar", methods=["GET"])
@login_required
def calendar():
    year = request.args.get("year", datetime.now().year, type=int)
    month = request.args.get("month", datetime.now().month, type=int)
    if month > 12:
        month = 1
        year += 1
    elif month < 1:
        month = 12
        year -= 1
    if current_user.role.role_name == "Admin":
        tasks = Task.query.filter(db.extract("year", Task.due_date) == year,
                                  db.extract("month", Task.due_date) == month).all()
    else:
        tasks = Task.query.filter(
            (Task.user_id == current_user.id) | (Task.assigned_to_id == current_user.id),
            db.extract("year", Task.due_date) == year,
            db.extract("month", Task.due_date) == month,
        ).all()
    weeks = monthcalendar(year, month)
    calendar_days = []
    for week in weeks:
        week_data = []
        for day in week:
            if day != 0:
                date = datetime(year, month, day)
                day_tasks = [t for t in tasks if t.due_date.date() == date.date()]
                week_data.append({"date": date, "tasks": day_tasks})
            else:
                week_data.append({"date": None, "tasks": []})
        calendar_days.append(week_data)
    prev_month = month - 1 if month > 1 else 12
    next_month = month + 1 if month < 12 else 1
    prev_year = year if month > 1 else year - 1
    next_year = year if month < 12 else year + 1
    easy = sum(1 for t in tasks if t.difficulty in ("1", "Легко"))
    medium = sum(1 for t in tasks if t.difficulty in ("2", "Средне"))
    hard = sum(1 for t in tasks if t.difficulty in ("3", "Сложно"))
    low_p = sum(1 for t in tasks if t.priority == "Низкий")
    med_p = sum(1 for t in tasks if t.priority == "Средний")
    high_p = sum(1 for t in tasks if t.priority == "Высокий")
    return render_template(
        "calendar.html",
        calendar_days=calendar_days,
        current_month_name=month_name[month],
        current_year=year,
        current_month=month,
        previous_month=prev_month,
        next_month=next_month,
        previous_year=prev_year,
        next_year=next_year,
        easy_count=easy,
        medium_count=medium,
        hard_count=hard,
        low_priority_count=low_p,
        medium_priority_count=med_p,
        high_priority_count=high_p,
    )


@app.before_request
def check_deadlines():
    if current_user.is_authenticated:
        in_progress = Task.query.filter_by(status="In Progress").all()
        for task in in_progress:
            if task.due_date < datetime.utcnow():
                task.status = "Overdue"
                db.session.commit()
        for task in Task.query.filter_by(user_id=current_user.id, status="In Progress").all():
            if task.due_date - datetime.utcnow() <= timedelta(days=2):
                flash(f'Задача "{task.title}" приближается к дедлайну!')


# ------------------ КАНБАН-ДОСКА (модифицированная с поддержкой сортировки по полезности) ------------------
@app.route("/task_board")
@login_required
def task_board():
    mode = request.args.get("mode", "status")
    project_id = request.args.get("project_id", type=int)
    sort_by = request.args.get("sort_by", "position")  # 'position' или 'utility'

    if project_id:
        project = Project.query.get_or_404(project_id)
        if current_user not in project.members and current_user.id != project.owner_id and current_user.role.role_name != "Admin":
            flash("Нет доступа к проекту", "error")
            return redirect(url_for("projects"))
        tasks = Task.query.filter_by(project_id=project_id).all()
    else:
        tasks = Task.query.filter((Task.user_id == current_user.id) | (Task.assigned_to_id == current_user.id)).all()

    # Если нужна сортировка по полезности
    if sort_by == 'utility':
        user_weights = get_user_weights(current_user.id)
        ranked = compute_normalized_utilities(tasks, user_weights)
        # Сортируем по убыванию полезности
        ranked.sort(key=lambda x: x[1], reverse=True)
        # Добавляем атрибут utility каждой задаче
        for task, utility, _, _, _ in ranked:
            task.utility = utility
        tasks = [task for task, _, _, _, _ in ranked]
    else:
        for task in tasks:
            task.utility = 0

    # Логика построения колонок
    if mode == "status":
        columns = KanbanColumn.query.filter(
            (KanbanColumn.project_id == project_id) | (KanbanColumn.project_id.is_(None))).order_by(
            KanbanColumn.order).all()
        tasks_by_column = {col.name: [] for col in columns}
        for task in tasks:
            col_name = task.status if task.status in tasks_by_column else (columns[0].name if columns else "To Do")
            tasks_by_column[col_name].append(task)
        for col in tasks_by_column:
            if sort_by == 'utility':
                tasks_by_column[col].sort(key=lambda t: getattr(t, 'utility', 0), reverse=True)
            else:
                tasks_by_column[col].sort(key=lambda t: t.position)
        column_counts = {col.name: len(tasks_by_column[col.name]) for col in columns}
    elif mode == "priority":
        columns = [{"name": "Низкий", "wip_limit": 0}, {"name": "Средний", "wip_limit": 0},
                   {"name": "Высокий", "wip_limit": 0}]
        tasks_by_column = {"Низкий": [], "Средний": [], "Высокий": []}
        for task in tasks:
            tasks_by_column[task.priority].append(task)
        for col in tasks_by_column:
            if sort_by == 'utility':
                tasks_by_column[col].sort(key=lambda t: getattr(t, 'utility', 0), reverse=True)
        column_counts = {col["name"]: len(tasks_by_column[col["name"]]) for col in columns}
    else:  # mode == "difficulty"
        columns = [{"name": "Легкая", "wip_limit": 0}, {"name": "Средняя", "wip_limit": 0},
                   {"name": "Сложная", "wip_limit": 0}]
        tasks_by_column = {"Легкая": [], "Средняя": [], "Сложная": []}
        for task in tasks:
            key = format_difficulty(task.difficulty)
            if key in tasks_by_column:
                tasks_by_column[key].append(task)
            else:
                tasks_by_column["Средняя"].append(task)
        for col in tasks_by_column:
            if sort_by == 'utility':
                tasks_by_column[col].sort(key=lambda t: getattr(t, 'utility', 0), reverse=True)
        column_counts = {col["name"]: len(tasks_by_column[col["name"]]) for col in columns}

    users = User.query.all()
    return render_template(
        "task_board.html",
        mode=mode,
        columns=columns,
        tasks_by_column=tasks_by_column,
        column_counts=column_counts,
        project_id=project_id,
        users=users,
        now=datetime.utcnow(),
        sort_by=sort_by
    )


@app.route("/task_board/move", methods=["POST"])
@login_required
def task_board_move():
    data = request.get_json()
    task_id = data.get("task_id")
    new_value = data.get("new_value")
    mode = data.get("mode")
    new_position = data.get("new_position")
    task = Task.query.get_or_404(task_id)
    if task.user_id != current_user.id and task.assigned_to_id != current_user.id and current_user.role.role_name != "Admin":
        return jsonify({"error": "Нет прав"}), 403
    if mode == "status":
        col = KanbanColumn.query.filter_by(name=new_value, project_id=task.project_id).first()
        if not col:
            col = KanbanColumn.query.filter_by(name=new_value, project_id=None).first()
        if col and col.wip_limit > 0:
            current_cnt = Task.query.filter(Task.status == new_value, Task.project_id == task.project_id).count()
            if current_cnt >= col.wip_limit:
                return jsonify({"error": f'WIP лимит {col.wip_limit} превышен в колонке "{new_value}"'}), 400
        now = datetime.utcnow()
        task.status = new_value
        if new_value == "In Progress" and not task.started_at:
            task.started_at = now
        if new_value == "Review" and not task.review_at:
            task.review_at = now
        if new_value == "Done":
            if not task.completed_at:
                task.completed_at = now
        else:
            task.completed_at = None
        history = task.status_history or []
        history.append({"status": new_value, "timestamp": now.isoformat()})
        task.status_history = history
    elif mode == "priority":
        task.priority = new_value
    else:
        if new_value == "Легкая":
            task.difficulty = "1"
        elif new_value == "Средняя":
            task.difficulty = "2"
        elif new_value == "Сложная":
            task.difficulty = "3"
        else:
            task.difficulty = new_value
    if new_position is not None and mode == "status":
        task.position = new_position
    db.session.commit()
    return jsonify({"success": True})


@app.route("/task_board/reorder", methods=["POST"])
@login_required
def task_board_reorder():
    data = request.get_json()
    col_name = data.get("column_name")
    task_ids = data.get("task_ids")
    if not col_name or not task_ids:
        return jsonify({"error": "Неверные данные"}), 400
    for idx, tid in enumerate(task_ids):
        t = Task.query.get(tid)
        if t and t.status == col_name:
            t.position = idx
    db.session.commit()
    return jsonify({"success": True})


# ------------------ ОТЧЁТЫ ------------------
def calculate_kanban_metrics(tasks):
    completed = [t for t in tasks if t.status == "Done" and t.completed_at]
    if not completed:
        return {"lead_time_avg": 0, "cycle_time_avg": 0, "throughput": 0, "total_completed": 0}
    lead_times = [(t.completed_at - t.created_at).total_seconds() / 86400 for t in completed if t.created_at]
    cycle_times = [(t.completed_at - t.started_at).total_seconds() / 86400 for t in completed if t.started_at]
    thirty_ago = datetime.utcnow() - timedelta(days=30)
    last30 = [t for t in completed if t.completed_at >= thirty_ago]
    throughput = len(last30) / 30.0
    return {
        "lead_time_avg": round(sum(lead_times) / len(lead_times), 2) if lead_times else 0,
        "cycle_time_avg": round(sum(cycle_times) / len(cycle_times), 2) if cycle_times else 0,
        "throughput": round(throughput, 2),
        "total_completed": len(completed),
    }


def get_completion_percentage(user_id=None):
    if user_id:
        tasks = Task.query.filter((Task.user_id == user_id) | (Task.assigned_to_id == user_id)).all()
    else:
        tasks = Task.query.all()
    total = len(tasks)
    if total == 0:
        return 0.0
    completed = sum(1 for t in tasks if t.status == "Done")
    return round((completed / total) * 100, 1)


@app.route("/reports", methods=["GET", "POST"])
@login_required
def reports():
    users = []
    if current_user.role.role_name == "Admin":
        users = User.query.all()
    if request.method == "POST" and current_user.role.role_name == "Admin" and request.form.get("user_id"):
        selected_user_id = int(request.form["user_id"])
        tasks = Task.query.filter((Task.user_id == selected_user_id) | (Task.assigned_to_id == selected_user_id)).all()
        user_kpi = get_completion_percentage(selected_user_id)
    else:
        tasks = Task.query.filter((Task.user_id == current_user.id) | (Task.assigned_to_id == current_user.id)).all()
        user_kpi = get_completion_percentage(current_user.id)
    kanban_metrics = calculate_kanban_metrics(tasks)
    company_kpi = get_completion_percentage()
    burn_data = {}
    for i in range(14):
        day = datetime.utcnow().date() - timedelta(days=i)
        cnt = Task.query.filter(
            db.func.date(Task.completed_at) == day,
            (Task.user_id == current_user.id) | (Task.assigned_to_id == current_user.id),
        ).count()
        burn_data[day.isoformat()] = cnt
    return render_template(
        "reports.html",
        user_kpi=user_kpi,
        company_kpi=company_kpi,
        kanban_metrics=kanban_metrics,
        burn_data=burn_data,
        users=users,
    )


@app.route("/download_report", methods=["POST"])
@login_required
def download_report():
    start = datetime.strptime(request.form["start_date"], "%Y-%m-%d")
    end = datetime.strptime(request.form["end_date"], "%Y-%m-%d")
    if current_user.role.role_name == "Admin" and request.form.get("user_id"):
        uid = int(request.form["user_id"])
    else:
        uid = current_user.id
    tasks = (
        Task.query.filter((Task.user_id == uid) | (Task.assigned_to_id == uid)).filter(
            Task.due_date.between(start, end)).all()
    )
    metrics = calculate_kanban_metrics(tasks)
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    headers = ["Название", "Описание", "Приоритет", "Статус", "Крайняя дата", "Сложность"]
    ws.append(headers)
    for t in tasks:
        ws.append([t.title, t.description, t.priority, t.status, t.due_date.strftime("%Y-%m-%d"), t.difficulty])
    ws.append([])
    ws.append(["Lead Time (средний, дни):", metrics["lead_time_avg"]])
    ws.append(["Cycle Time (средний, дни):", metrics["cycle_time_avg"]])
    ws.append(["Throughput (задач/день):", metrics["throughput"]])
    ws.append(["Всего завершено:", metrics["total_completed"]])
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f'report_{start.strftime("%Y-%m-%d")}_to_{end.strftime("%Y-%m-%d")}.xlsx',
    )


@app.route("/download_all_kpis", methods=["POST"])
@login_required
def download_all_kpis():
    if current_user.role.role_name != "Admin":
        flash("Нет прав")
        return redirect(url_for("reports"))
    start = datetime.strptime(request.form["start_date"], "%Y-%m-%d")
    end = datetime.strptime(request.form["end_date"], "%Y-%m-%d")
    users = User.query.all()
    data = []
    for u in users:
        tasks = (
            Task.query.filter((Task.user_id == u.id) | (Task.assigned_to_id == u.id))
            .filter(Task.due_date.between(start, end))
            .all()
        )
        m = calculate_kanban_metrics(tasks)
        data.append(
            {"user": u.username, "lead": m["lead_time_avg"], "cycle": m["cycle_time_avg"],
             "throughput": m["throughput"]}
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "Общий KPI"
    ws.append(["Сотрудник", "Lead Time (дни)", "Cycle Time (дни)", "Throughput (задач/день)"])
    for d in data:
        ws.append([d["user"], d["lead"], d["cycle"], d["throughput"]])
    from openpyxl.chart import BarChart, Reference

    chart = BarChart()
    chart.title = "Lead Time сотрудников"
    chart.x_axis.title = "Сотрудники"
    chart.y_axis.title = "Дни"
    data_ref = Reference(ws, min_col=2, min_row=2, max_row=len(data) + 1, max_col=2)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(data) + 1)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, "E4")
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out, as_attachment=True,
        download_name=f'all_kpis_{start.strftime("%Y-%m-%d")}_to_{end.strftime("%Y-%m-%d")}.xlsx'
    )


# ------------------ ДИАГРАММА ГАНТА, ПОДЗАДАЧИ, КОММЕНТАРИИ ------------------
@app.route("/gantt", methods=["GET", "POST"])
@login_required
def gantt():
    if request.method == "POST":
        new = Task(
            title=request.form["title"],
            description=request.form["description"],
            due_date=datetime.strptime(request.form["due_date"], "%Y-%m-%d"),
            priority=request.form["priority"],
            status=request.form["status"],
            difficulty=request.form["difficulty"],
            user_id=current_user.id,
            assigned_to_id=request.form.get("assigned_to"),
        )
        db.session.add(new)
        db.session.commit()
        return redirect(url_for("gantt"))
    tasks = Task.query.all()
    return render_template("gantt.html", tasks_data=[t.to_dict() for t in tasks], users=User.query.all())


@app.route("/add_subtask/<int:task_id>", methods=["POST"])
@login_required
def add_subtask(task_id):
    data = request.get_json()
    st = SubTask(
        task_id=task_id,
        title=data["title"],
        start_date=datetime.strptime(data["start_date"], "%Y-%m-%d"),
        end_date=datetime.strptime(data["end_date"], "%Y-%m-%d"),
    )
    db.session.add(st)
    db.session.commit()
    return jsonify({"status": "success", "subtask": st.to_dict()})


@app.route("/delete_subtask/<int:subtask_id>", methods=["DELETE"])
@login_required
def delete_subtask(subtask_id):
    st = SubTask.query.get(subtask_id)
    if st:
        db.session.delete(st)
        db.session.commit()
        return jsonify({"status": "success"})
    return jsonify({"status": "error"}), 404


@app.route("/create_subtask", methods=["POST"])
@login_required
def create_subtask():
    data = request.get_json()
    st = SubTask(
        task_id=data["task_id"],
        title=data["title"],
        start_date=datetime.strptime(data["start_date"], "%Y-%m-%d"),
        end_date=datetime.strptime(data["end_date"], "%Y-%m-%d"),
    )
    db.session.add(st)
    db.session.commit()
    return jsonify({"status": "success", "subtask": st.to_dict()})


@app.route("/update_subtask", methods=["POST"])
@login_required
def update_subtask():
    data = request.get_json()
    st = SubTask.query.get(data["id"])
    if st:
        st.title = data["title"]
        st.start_date = datetime.strptime(data["start_date"], "%Y-%m-%d")
        st.end_date = datetime.strptime(data["end_date"], "%Y-%m-%d")
        db.session.commit()
        return jsonify({"status": "success", "subtask": st.to_dict()})
    return jsonify({"status": "error"}), 404


@app.route("/task/<int:task_id>/add_comment", methods=["POST"])
@login_required
def add_comment(task_id):
    data = request.get_json()
    if not data.get("comment", "").strip():
        return jsonify({"status": "error", "message": "Пустой комментарий"}), 400
    comm = TaskComments(task_id=task_id, user_id=current_user.id, comment=data["comment"])
    db.session.add(comm)
    db.session.commit()
    return jsonify({"status": "success", "comment": comm.to_dict()})


@app.route("/task/<int:task_id>/edit_comment/<int:comment_id>", methods=["PUT"])
@login_required
def edit_comment(task_id, comment_id):
    data = request.get_json()
    comm = TaskComments.query.filter_by(id=comment_id, task_id=task_id).first()
    if not comm:
        return jsonify({"status": "error"}), 404
    if comm.user_id != current_user.id:
        return jsonify({"status": "error"}), 403
    comm.comment = data["comment"]
    db.session.commit()
    return jsonify({"status": "success", "comment": comm.to_dict()})


@app.route("/task/<int:task_id>/delete_comment/<int:comment_id>", methods=["DELETE"])
@login_required
def delete_comment(task_id, comment_id):
    comm = TaskComments.query.filter_by(id=comment_id, task_id=task_id).first()
    if not comm:
        return jsonify({"status": "error"}), 404
    if comm.user_id != current_user.id:
        return jsonify({"status": "error"}), 403
    db.session.delete(comm)
    db.session.commit()
    return jsonify({"status": "success"})


@app.route("/task/<int:task_id>")
@login_required
def get_task(task_id):
    task = Task.query.get_or_404(task_id)
    return jsonify({"id": task.id, "title": task.title, "comments": [c.to_dict() for c in task.task_comments]})


# ------------------ РЕДАКТОР ДОКУМЕНТОВ ------------------
@app.route("/editor")
def editor_page():
    files = [f for f in os.listdir(UPLOAD_FOLDER) if os.path.isfile(os.path.join(UPLOAD_FOLDER, f))]
    return render_template("editor.html", files=files)


@app.route("/edit/<filename>")
def edit_document(filename):
    key = md5(f"{filename}_{current_user.id}_{datetime.utcnow()}".encode()).hexdigest()
    config = {
        "document": {
            "fileType": filename.split(".")[-1].lower(),
            "key": key,
            "title": filename,
            "url": f"http://127.0.0.1:5000/uploads/{filename}",
            "permissions": {"edit": True},
        },
        "editorConfig": {
            "callbackUrl": "http://127.0.0.1:5000/callback",
            "mode": "edit",
            "lang": "ru",
            "user": {"id": current_user.id, "name": current_user.username},
        },
    }
    return render_template("editor_frame.html", document_config=config)


@app.route("/upload/<task_id>", methods=["POST"])
def upload_file_for_task(task_id):
    if "file" not in request.files:
        return "No file", 400
    f = request.files["file"]
    if f.filename == "":
        return "No file selected", 400
    if not allowed_file(f.filename):
        return "Недопустимый формат", 400
    f.save(os.path.join(UPLOAD_FOLDER, secure_filename(f.filename)))
    db.session.add(File(filename=f.filename, task_id=task_id))
    db.session.commit()
    return redirect(url_for("editor_page"))


@app.route("/callback", methods=["POST"])
def callback():
    app.logger.info(f"Callback: {request.json}")
    return jsonify({"error": 0})


# ------------------ НАПОМИНАНИЯ ------------------
@app.route("/reminders")
@login_required
def view_reminders():
    rems = Reminder.query.filter_by(user_id=current_user.id).order_by(Reminder.reminder_date).all()
    return render_template("reminders.html", reminders=rems)


@app.route("/reminder/new", methods=["GET", "POST"])
@login_required
def create_reminder():
    if request.method == "POST":
        r = Reminder(
            title=request.form["title"],
            description=request.form.get("description", ""),
            reminder_date=datetime.strptime(request.form["reminder_date"], "%Y-%m-%dT%H:%M"),
            priority=request.form["priority"],
            repeat=request.form["repeat"],
            user_id=current_user.id,
        )
        db.session.add(r)
        db.session.commit()
        flash("Напоминание создано!", "success")
        return redirect(url_for("view_reminders"))
    return render_template("create_reminder.html")


@app.route("/reminder/edit/<int:reminder_id>", methods=["GET", "POST"])
@login_required
def edit_reminder(reminder_id):
    r = Reminder.query.filter_by(id=reminder_id, user_id=current_user.id).first_or_404()
    if request.method == "POST":
        r.title = request.form["title"]
        r.description = request.form.get("description", "")
        r.reminder_date = datetime.strptime(request.form["reminder_date"], "%Y-%m-%dT%H:%M")
        r.priority = request.form["priority"]
        r.repeat = request.form["repeat"]
        db.session.commit()
        flash("Обновлено!", "success")
        return redirect(url_for("view_reminders"))
    return render_template("edit_reminder.html", reminder=r)


@app.route("/reminder/delete/<int:reminder_id>", methods=["POST"])
@login_required
def delete_reminder(reminder_id):
    r = Reminder.query.filter_by(id=reminder_id, user_id=current_user.id).first_or_404()
    db.session.delete(r)
    db.session.commit()
    flash("Удалено!", "success")
    return redirect(url_for("view_reminders"))


# ------------------ ПРОЕКТЫ ------------------
@app.route("/projects")
@login_required
def projects():
    projs = Project.query.filter(
        (Project.owner_id == current_user.id) | (Project.members.any(id=current_user.id))).all()
    return render_template("projects.html", projects=projs)


@app.route("/project/new", methods=["GET", "POST"])
@login_required
def create_project():
    if request.method == "POST":
        p = Project(name=request.form["name"], description=request.form.get("description", ""),
                    owner_id=current_user.id)
        db.session.add(p)
        db.session.commit()
        flash("Проект создан!", "success")
        return redirect(url_for("projects"))
    return render_template("create_project.html")


@app.route("/project/<int:project_id>")
def project_detail(project_id):
    project = Project.query.get_or_404(project_id)
    return render_template("project_detail.html", project=project, all_users=User.query.all())


@app.route("/project/<int:project_id>/add_member", methods=["POST"])
def add_project_member(project_id):
    p = Project.query.get_or_404(project_id)
    if p.owner_id != current_user.id:
        flash("Только владелец может добавлять участников", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    user = User.query.get(request.form["user_id"])
    if user:
        p.members.append(user)
        db.session.commit()
        flash("Участник добавлен", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.route("/project/<int:project_id>/task/new", methods=["GET", "POST"])
@login_required
def create_task_in_project(project_id):
    p = Project.query.get_or_404(project_id)
    if current_user.id != p.owner_id and current_user not in p.members:
        flash("Нет доступа", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    if request.method == "POST":
        new = Task(
            title=request.form["title"],
            description=request.form["description"],
            priority=request.form["priority"],
            status=request.form["status"],
            difficulty=request.form["difficulty"],
            due_date=datetime.strptime(request.form["due_date"], "%Y-%m-%d"),
            project_id=project_id,
            user_id=current_user.id,
            assigned_to_id=request.form.get("assigned_to"),
            position=0,
        )
        db.session.add(new)
        db.session.commit()
        flash("Задача добавлена!", "success")
        return redirect(url_for("project_detail", project_id=project_id))
    columns = (
        KanbanColumn.query.filter((KanbanColumn.project_id == project_id) | (KanbanColumn.project_id.is_(None)))
        .order_by(KanbanColumn.order)
        .all()
    )
    status_choices = [(c.name, c.name) for c in columns]
    return render_template("create_task.html", project=p, users=p.members, status_choices=status_choices)


def project_access_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        p = Project.query.get_or_404(kwargs["project_id"])
        if current_user.id != p.owner_id and current_user not in p.members:
            flash("Нет доступа", "error")
            return redirect(url_for("projects"))
        return func(*args, **kwargs)

    return wrapper


@app.route("/project/<int:project_id>/delete", methods=["POST"])
@login_required
def delete_project(project_id):
    p = Project.query.get_or_404(project_id)
    if p.owner_id != current_user.id:
        flash("Только владелец может удалить проект", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    db.session.delete(p)
    db.session.commit()
    flash("Проект удалён", "success")
    return redirect(url_for("projects"))


@app.route("/project/<int:project_id>/remove_member", methods=["POST"])
@login_required
def remove_project_member(project_id):
    p = Project.query.get_or_404(project_id)
    if p.owner_id != current_user.id:
        flash("Только владелец может удалять участников", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    user = User.query.get(request.form["user_id"])
    if user and user in p.members:
        p.members.remove(user)
        db.session.commit()
        flash("Участник удалён", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.route("/project/<int:project_id>/edit", methods=["GET", "POST"])
@login_required
def edit_project(project_id):
    p = Project.query.get_or_404(project_id)
    if p.owner_id != current_user.id:
        flash("Только владелец может редактировать", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    if request.method == "POST":
        p.name = request.form["name"]
        p.description = request.form.get("description", "")
        db.session.commit()
        flash("Проект обновлён", "success")
        return redirect(url_for("project_detail", project_id=project_id))
    return render_template("edit_project.html", project=p)


@app.route("/project_tasks/<int:project_id>")
@login_required
def project_tasks(project_id):
    return redirect(url_for("task_board", project_id=project_id))


@app.route("/project_tasks_priority/<int:project_id>")
@login_required
def project_tasks_priority(project_id):
    return redirect(url_for("task_board", project_id=project_id, mode="priority"))


@app.route("/project_tasks_difficulty/<int:project_id>")
@login_required
def project_tasks_difficulty(project_id):
    return redirect(url_for("task_board", project_id=project_id, mode="difficulty"))


@app.route("/update_project_task_category", methods=["POST"])
@login_required
def update_project_task_category():
    data = request.json
    task = Task.query.get(data.get("task_id"))
    if task:
        status_map = {"В работе": "In Progress", "Выполненные": "Done", "Просроченные": "Overdue"}
        new_status = status_map.get(data.get("new_category"), data.get("new_category"))
        task.status = new_status
        db.session.commit()
        return jsonify({"success": True})
    return jsonify({"error": "Задача не найдена"}), 404


# ------------------ АДМИНИСТРИРОВАНИЕ ПОЛЬЗОВАТЕЛЕЙ ------------------
@app.route("/admin/users")
@login_required
def manage_users():
    if current_user.role.role_name != "Admin":
        flash("Нет прав", "error")
        return redirect(url_for("dashboard"))
    return render_template("admin_users.html", users=User.query.all(), roles=Role.query.all())


@app.route("/admin_user/<int:user_id>/update_role", methods=["POST"])
@login_required
def update_user_role(user_id):
    if current_user.role.role_name != "Admin":
        return jsonify({"status": "error"}), 403
    user = User.query.get_or_404(user_id)
    data = request.get_json()
    user.role_id = data.get("role_id")
    db.session.commit()
    return jsonify({"status": "success"})


@app.route("/admin/user/<int:user_id>/delete", methods=["POST"])
@login_required
def delete_user(user_id):
    if current_user.role.role_name != "Admin":
        flash("Нет прав", "error")
        return redirect(url_for("manage_users"))
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash("Нельзя удалить себя", "error")
        return redirect(url_for("manage_users"))
    db.session.delete(user)
    db.session.commit()
    flash(f"Пользователь {user.username} удалён", "success")
    return redirect(url_for("manage_users"))


@app.route("/admin_user/<int:user_id>/update_password", methods=["POST"])
@login_required
def update_password(user_id):
    data = request.json
    user = User.query.get(user_id)
    if user:
        user.password = data.get("password")
        db.session.commit()
        return jsonify({"status": "success"})
    return jsonify({"status": "error"}), 404


@app.route("/admin_user/<int:user_id>/update_name", methods=["POST"])
@login_required
def update_username(user_id):
    if current_user.role.role_name != "Admin":
        return jsonify({"status": "error"}), 403
    data = request.get_json()
    user = User.query.get_or_404(user_id)
    user.username = data.get("username")
    db.session.commit()
    return jsonify({"status": "success"})


@app.route("/admin/users/create", methods=["GET", "POST"])
@login_required
def create_user_admin():
    if current_user.role.role_name != "Admin":
        flash("Нет прав", "error")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form["username"]
        if User.query.filter_by(username=username).first():
            flash("Имя уже существует", "error")
            return redirect(url_for("create_user_admin"))
        new = User(
            username=username,
            password=request.form["password"],
            role_id=request.form["role_id"],
            email=request.form.get("email"),
        )
        db.session.add(new)
        db.session.commit()
        flash("Пользователь создан", "success")
        return redirect(url_for("manage_users"))
    return render_template("admin_create_user.html", roles=Role.query.all())


# ================= [PROCESS MINING] МАРШРУТЫ =================
@app.route('/process_mining')
@login_required
def process_mining():
    if current_user.role.role_name == 'Admin':
        events = get_task_event_log()
    else:
        events = get_task_event_log(user_id=current_user.id)

    metrics = compute_process_metrics(events)
    transitions = metrics['transitions']

    regress_tasks = set()
    for ev in events:
        if ev.get('from_status') == 'Review' and ev.get('to_status') == 'In Progress':
            regress_tasks.add(ev['title'])
    regress_count = len(regress_tasks)

    recommendations = generate_recommendations(metrics, transitions, regress_count)

    return render_template('process_mining.html',
                           metrics=metrics,
                           transitions=transitions,
                           regress_count=regress_count,
                           recommendations=recommendations)


@app.route('/export_event_log')
@login_required
def export_event_log():
    import csv
    from io import StringIO

    if current_user.role.role_name == 'Admin':
        events = get_task_event_log()
    else:
        events = get_task_event_log(user_id=current_user.id)

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['task_id', 'title', 'timestamp', 'from_status', 'to_status', 'user'])
    for ev in events:
        writer.writerow([ev['task_id'], ev['title'], ev['timestamp'], ev['from_status'], ev['to_status'], ev['user']])

    response = app.response_class(
        response=output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=event_log.csv'}
    )
    return response


@app.route('/process_mining_graph.png')
@login_required
def process_mining_graph():
    if current_user.role.role_name == 'Admin':
        events = get_task_event_log()
    else:
        events = get_task_event_log(user_id=current_user.id)
    metrics = compute_process_metrics(events)
    transitions = metrics['transitions']

    img_path = generate_transition_graph(transitions)
    if img_path is None:
        return '', 204
    return send_file(img_path, mimetype='image/png', as_attachment=False)


@app.route('/api/process_metrics')
@login_required
def api_process_metrics():
    if current_user.role.role_name == 'Admin':
        events = get_task_event_log()
        tasks = Task.query.all()
    else:
        events = get_task_event_log(user_id=current_user.id)
        tasks = Task.query.filter((Task.user_id == current_user.id) | (Task.assigned_to_id == current_user.id)).all()

    metrics = compute_process_metrics(events)
    extended = calculate_extended_metrics(events)

    # Добавляем среднюю полезность и долю высокополезных задач
    user_weights = get_user_weights(current_user.id)
    ranked = compute_normalized_utilities(tasks, user_weights)
    avg_utility = sum(util for _, util, _, _, _ in ranked) / len(ranked) if ranked else 0
    high_utility_count = sum(1 for _, util, _, _, _ in ranked if util > 0.7)

    return jsonify({
        'status_durations': metrics['avg_duration_per_status'],
        'total_tasks': metrics['total_tasks'],
        'total_events': metrics['total_events'],
        'lead_time_avg': extended['avg_lead_time_hours'],
        'cycle_time_avg': extended['avg_cycle_time_hours'],
        'throughput': extended['throughput_per_day'],
        'completed_tasks': extended['completed_tasks'],
        'performer_stats': extended['performer_stats'],
        'avg_utility': round(avg_utility, 4),
        'high_utility_tasks_ratio': round(high_utility_count / len(tasks) if tasks else 0, 4)
    })


@app.route('/api/transition_sankey')
@login_required
def api_transition_sankey():
    if current_user.role.role_name == 'Admin':
        events = get_task_event_log()
    else:
        events = get_task_event_log(user_id=current_user.id)
    metrics = compute_process_metrics(events)
    sankey = get_sankey_data(metrics['transitions'])
    return jsonify(sankey)


@app.route('/api/timeline_heatmap')
@login_required
def api_timeline_heatmap():
    if current_user.role.role_name == 'Admin':
        events = get_task_event_log()
    else:
        events = get_task_event_log(user_id=current_user.id)
    timeline = get_timeline_events(events)
    return jsonify(timeline)


# ================= НОВЫЕ API ДЛЯ УПРАВЛЕНИЯ ВЕСАМИ И РАНЖИРОВАНИЯ =================
@app.route('/api/task_ranking')
@login_required
def api_task_ranking():
    """Возвращает задачи пользователя (или проекта) с рассчитанной полезностью."""
    project_id = request.args.get('project_id', type=int)
    query = Task.query
    if project_id:
        project = Project.query.get_or_404(project_id)
        if current_user not in project.members and current_user.id != project.owner_id and current_user.role.role_name != 'Admin':
            return jsonify({'error': 'Нет доступа'}), 403
        query = query.filter_by(project_id=project_id)
    else:
        query = query.filter((Task.user_id == current_user.id) | (Task.assigned_to_id == current_user.id))

    tasks = query.all()
    user_weights = get_user_weights(current_user.id)
    ranked = compute_normalized_utilities(tasks, user_weights)
    ranked.sort(key=lambda x: x[1], reverse=True)
    result = []
    for task, utility, norm_p, norm_d, norm_t in ranked:
        result.append({
            'id': task.id,
            'title': task.title,
            'status': task.status,
            'priority': task.priority,
            'difficulty': task.difficulty,
            'utility': round(utility, 4),
            'normalized_priority': round(norm_p, 4),
            'normalized_difficulty': round(norm_d, 4),
            'normalized_time': round(norm_t, 4)
        })
    return jsonify(result)


@app.route('/api/set_weights', methods=['POST'])
@login_required
def set_weights():
    """Установка весов критериев вручную."""
    data = request.get_json()
    w_pri = float(data.get('weight_priority', 0.5))
    w_diff = float(data.get('weight_difficulty', 0.3))
    w_time = float(data.get('weight_time', 0.2))
    total = w_pri + w_diff + w_time
    if total > 0:
        w_pri /= total
        w_diff /= total
        w_time /= total
    user_weights = get_user_weights(current_user.id)
    user_weights.weight_priority = w_pri
    user_weights.weight_difficulty = w_diff
    user_weights.weight_time = w_time
    db.session.commit()
    return jsonify({'status': 'success', 'weights': {'priority': w_pri, 'difficulty': w_diff, 'time': w_time}})


@app.route('/api/set_weights_by_ahp', methods=['POST'])
@login_required
def set_weights_by_ahp():
    """Установка весов через матрицу парных сравнений (МАИ)."""
    data = request.get_json()
    matrix = data.get('matrix')
    if not matrix or len(matrix) != 3 or any(len(row) != 3 for row in matrix):
        return jsonify({'error': 'Матрица должна быть 3x3'}), 400
    weights, cr = ahp_weights_from_matrix(matrix)
    user_weights = get_user_weights(current_user.id)
    user_weights.weight_priority = weights[0]
    user_weights.weight_difficulty = weights[1]
    user_weights.weight_time = weights[2]
    user_weights.ahp_matrix = matrix
    db.session.commit()
    return jsonify({
        'status': 'success',
        'weights': {'priority': weights[0], 'difficulty': weights[1], 'time': weights[2]},
        'consistency_ratio': cr
    })


@app.route('/api/get_weights')
@login_required
def get_weights():
    """Возвращает текущие веса пользователя."""
    uw = get_user_weights(current_user.id)
    return jsonify({
        'priority': uw.weight_priority,
        'difficulty': uw.weight_difficulty,
        'time': uw.weight_time
    })

@app.route('/weights_settings')
@login_required
def weights_settings():
    """Страница настройки весов критериев (МАИ и ручная)."""
    return render_template('weights_settings.html')

# ------------------ ИНИЦИАЛИЗАЦИЯ КОЛОНОК КАНБАН И ВЕСОВ ------------------
def init_kanban_columns():
    default = [
        {"name": "To Do", "wip_limit": 5, "order": 0},
        {"name": "In Progress", "wip_limit": 3, "order": 1},
        {"name": "Review", "wip_limit": 2, "order": 2},
        {"name": "Done", "wip_limit": 0, "order": 3},
    ]
    for col in default:
        if not KanbanColumn.query.filter_by(name=col["name"], project_id=None).first():
            db.session.add(
                KanbanColumn(name=col["name"], wip_limit=col["wip_limit"], order=col["order"], project_id=None))
    db.session.commit()


def init_database():
    with app.app_context():
        db.create_all()
        init_kanban_columns()
        # Создаём записи весов для существующих пользователей (если нет)
        for user in User.query.all():
            if not UserWeight.query.filter_by(user_id=user.id).first():
                db.session.add(UserWeight(user_id=user.id))
        db.session.commit()


if __name__ == "__main__":
    init_database()
    app.run(debug=True)